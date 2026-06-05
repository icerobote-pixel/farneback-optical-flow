from __future__ import annotations

import csv
import math
from pathlib import Path
from typing import Dict, Tuple

import cv2
import numpy as np

from .save_manager import safe_imwrite


def draw_quiver(frame_bgr, flow, step=16, scale=2.0, show_text=False, mag_text_th=0.15):
    h, w = frame_bgr.shape[:2]
    vis = frame_bgr.copy()

    ys, xs = np.mgrid[step // 2:h:step, step // 2:w:step].astype(np.int32)
    fx = flow[ys, xs, 0]
    fy = flow[ys, xs, 1]

    fxv = fx.reshape(-1)
    fyv = fy.reshape(-1)
    pts0 = np.stack([xs, ys], axis=-1).reshape(-1, 2)
    pts1 = np.stack(
        [xs + (fx * scale).astype(np.int32), ys + (fy * scale).astype(np.int32)],
        axis=-1,
    ).reshape(-1, 2)

    for (x0, y0), (x1, y1), dx, dy in zip(pts0, pts1, fxv, fyv):
        cv2.arrowedLine(vis, (int(x0), int(y0)), (int(x1), int(y1)), (0, 255, 0), 1, tipLength=0.3)
        if show_text:
            mag = float(np.sqrt(dx * dx + dy * dy))
            if mag >= mag_text_th:
                ang = (math.degrees(math.atan2(dy, dx)) + 360.0) % 360.0
                text = f"{ang:.0f}|{mag:.2f}"
                cv2.putText(
                    vis,
                    text,
                    (int(x0) + 2, int(y0) - 2),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.33,
                    (0, 0, 255),
                    1,
                    cv2.LINE_AA,
                )
    return vis


def flow_to_hsv(flow):
    fx, fy = flow[..., 0], flow[..., 1]
    mag, ang = cv2.cartToPolar(fx, fy, angleInDegrees=True)
    h = (ang / 2).astype(np.uint8)
    s = np.full_like(h, 255, dtype=np.uint8)
    max_mag = float(mag.max()) if np.isfinite(mag).any() and float(mag.max()) > 0 else 1.0
    v = np.clip((mag / (max_mag + 1e-6) * 255), 0, 255).astype(np.uint8)
    hsv = cv2.merge([h, s, v])
    return cv2.cvtColor(hsv, cv2.COLOR_HSV2BGR)


def mean_flow_stats(flow):
    fx = flow[..., 0]
    fy = flow[..., 1]
    fx = fx[np.isfinite(fx)]
    fy = fy[np.isfinite(fy)]
    nonzero = (fx != 0) | (fy != 0)
    fx = fx[nonzero]
    fy = fy[nonzero]
    if fx.size == 0:
        return 0.0, 0.0
    mean_fx = float(np.mean(fx))
    mean_fy = float(np.mean(fy))
    speed = float(np.sqrt(mean_fx ** 2 + mean_fy ** 2))
    angle_deg = (math.degrees(math.atan2(mean_fy, mean_fx)) + 360.0) % 360.0
    return speed, angle_deg


def bool_mask_to_u8(mask_bool):
    if mask_bool is None:
        return None
    return mask_bool.astype(np.uint8) * 255


def overlay_mask(frame_bgr, mask_bool, color=(0, 0, 255), alpha=0.35):
    vis = frame_bgr.copy()
    if mask_bool is None:
        return vis
    mask_bool = mask_bool.astype(bool)
    if not mask_bool.any():
        return vis
    color_layer = np.zeros_like(frame_bgr, dtype=np.uint8)
    color_layer[:] = color
    vis[mask_bool] = cv2.addWeighted(frame_bgr[mask_bool], 1.0 - alpha, color_layer[mask_bool], alpha, 0)
    return vis


def make_region_id_map(shape_hw: Tuple[int, int], contours):
    h, w = shape_hw
    rid = np.zeros((h, w), dtype=np.int32)
    for idx, cnt in enumerate(contours, start=1):
        cv2.drawContours(rid, [cnt], -1, idx, thickness=-1)
    return rid


def postprocess_target_mask(mask_bool, post_cfg: Dict):
    """对 two-stage 得到的 mask 做形态学和面积后处理。"""
    if mask_bool is None:
        return dict(
            pre_mask_bool=None,
            morph_mask_bool=None,
            post_mask_bool=None,
            pre_mask_u8=None,
            morph_mask_u8=None,
            post_mask_u8=None,
            raw_contours=[],
            kept_contours=[],
            kept_boxes=[],
            region_id_map=None,
            num_raw_contours=0,
            num_kept_contours=0,
            pre_target_pixels=0,
            morph_target_pixels=0,
            post_target_pixels=0,
        )

    pre_mask_bool = mask_bool.astype(bool)
    pre_mask_u8 = bool_mask_to_u8(pre_mask_bool)

    if post_cfg["ENABLE_MASK_MORPHOLOGY"]:
        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, post_cfg["MORPH_KERNEL_SIZE"])
        morph_mask_u8 = pre_mask_u8.copy()
        if post_cfg["MORPH_OPEN_ITER"] > 0:
            morph_mask_u8 = cv2.morphologyEx(
                morph_mask_u8, cv2.MORPH_OPEN, kernel, iterations=post_cfg["MORPH_OPEN_ITER"]
            )
        if post_cfg["MORPH_CLOSE_ITER"] > 0:
            morph_mask_u8 = cv2.morphologyEx(
                morph_mask_u8, cv2.MORPH_CLOSE, kernel, iterations=post_cfg["MORPH_CLOSE_ITER"]
            )
    else:
        morph_mask_u8 = pre_mask_u8.copy()

    morph_mask_bool = morph_mask_u8 > 0
    contours, _ = cv2.findContours(morph_mask_u8, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    post_mask_u8 = np.zeros_like(morph_mask_u8)
    kept_contours = []
    kept_boxes = []
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if post_cfg["ENABLE_AREA_FILTER"] and (
            area < post_cfg["MIN_REGION_AREA"] or area > post_cfg["MAX_REGION_AREA"]
        ):
            continue
        kept_contours.append(cnt)
        x, y, bw, bh = cv2.boundingRect(cnt)
        kept_boxes.append((x, y, bw, bh))
        cv2.drawContours(post_mask_u8, [cnt], -1, 255, thickness=-1)

    post_mask_bool = post_mask_u8 > 0
    region_id_map = make_region_id_map(post_mask_bool.shape, kept_contours)

    return dict(
        pre_mask_bool=pre_mask_bool,
        morph_mask_bool=morph_mask_bool,
        post_mask_bool=post_mask_bool,
        pre_mask_u8=pre_mask_u8,
        morph_mask_u8=morph_mask_u8,
        post_mask_u8=post_mask_u8,
        raw_contours=contours,
        kept_contours=kept_contours,
        kept_boxes=kept_boxes,
        region_id_map=region_id_map,
        num_raw_contours=len(contours),
        num_kept_contours=len(kept_contours),
        pre_target_pixels=int(pre_mask_bool.sum()),
        morph_target_pixels=int(morph_mask_bool.sum()),
        post_target_pixels=int(post_mask_bool.sum()),
    )


def draw_post_frames(frame_bgr, post_pack):
    region_frame = frame_bgr.copy()
    box_frame = frame_bgr.copy()
    combined_frame = frame_bgr.copy()

    for cnt, (x, y, bw, bh) in zip(post_pack["kept_contours"], post_pack["kept_boxes"]):
        cv2.drawContours(region_frame, [cnt], -1, (0, 0, 255), 2)
        cv2.drawContours(combined_frame, [cnt], -1, (0, 0, 255), 2)

        cx = x + bw / 2.0
        cy = y + bh / 2.0
        text = f"cx={int(cx)}, cy={int(cy)}, w={bw}, h={bh}"

        cv2.rectangle(box_frame, (x, y), (x + bw, y + bh), (0, 255, 255), 2)
        cv2.rectangle(combined_frame, (x, y), (x + bw, y + bh), (0, 255, 255), 2)
        cv2.circle(box_frame, (int(cx), int(cy)), 3, (255, 0, 0), -1)
        cv2.circle(combined_frame, (int(cx), int(cy)), 3, (255, 0, 0), -1)

        ty = max(0, y - 5)
        pos = (x, ty if ty > 0 else y + 15)
        cv2.putText(box_frame, text, pos, cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 255, 255), 1, cv2.LINE_AA)
        cv2.putText(combined_frame, text, pos, cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 255, 255), 1, cv2.LINE_AA)

    return region_frame, box_frame, combined_frame


def save_debug(debug_pack, frame_idx: int, vis_cfg: Dict, paths: Dict, logger=None):
    if not vis_cfg["SAVE_DEBUG_PLOTS"] and not vis_cfg["SAVE_DEBUG_TABLE_CSV"]:
        return

    try:
        import matplotlib.pyplot as plt

        pts = debug_pack["pts"]
        mag = debug_pack["mag"]
        theta = debug_pack["theta"]
        xs = debug_pack["xs"]
        ys = debug_pack["ys"]
        dom_dir = debug_pack["dom_dir"]
        hist = debug_pack["hist"]
        dtheta_deg = debug_pack["dtheta_deg"]
        r_theta = debug_pack["r_theta"]
        r0_theta = debug_pack["r0_theta"]
        mag0 = debug_pack["mag0"]
        r_mag = debug_pack["r_mag"]
        r0_mag = debug_pack["r0_mag"]
        inside_dir = debug_pack["inside_dir"].astype(bool)
        mag_outlier = debug_pack["mag_outlier"].astype(bool)
        target_s = debug_pack["target_s"].astype(bool)

        out_dir = paths["cluster_plot_dir"]

        if vis_cfg["SAVE_DEBUG_PLOTS"]:
            fig, ax = plt.subplots(figsize=(5.2, 5.2))
            ax.scatter(pts[~target_s, 0], pts[~target_s, 1], s=4, alpha=0.25, label="background-like")
            ax.scatter(pts[target_s, 0], pts[target_s, 1], s=7, alpha=0.75, label="target")
            ax.set_xlabel("fx")
            ax.set_ylabel("fy")
            ax.set_title(f"Frame {frame_idx} | two-stage target vs bg")
            ax.grid(True)
            ax.legend(loc="best")
            fig.tight_layout()
            fig.savefig(str(out_dir / f"scatter_fx_fy_frame_{frame_idx:05d}.png"), dpi=150)
            plt.close(fig)

            fig, ax = plt.subplots(figsize=(7.2, 3.2))
            ax.plot(hist, linewidth=1.5)
            ax.set_title(f"Frame {frame_idx} angle histogram")
            ax.set_xlabel("angle bin")
            ax.set_ylabel("count")
            ax.grid(True)
            fig.tight_layout()
            fig.savefig(str(out_dir / f"angle_hist_frame_{frame_idx:05d}.png"), dpi=150)
            plt.close(fig)

            d_sorted = np.sort(dtheta_deg)
            fig, ax = plt.subplots(figsize=(6.8, 3.8))
            ax.plot(np.arange(d_sorted.size), d_sorted, linewidth=1.5)
            ax.axhline(r0_theta, linestyle="--", linewidth=1, label=f"r0_theta={r0_theta:.2f} deg")
            ax.axhline(r_theta, linestyle="--", linewidth=1, label=f"r_theta={r_theta:.2f} deg")
            ax.set_title(f"Frame {frame_idx} knee on |Δθ|")
            ax.set_xlabel("sorted index")
            ax.set_ylabel("|Δθ| (deg)")
            ax.grid(True)
            ax.legend(loc="best")
            fig.tight_layout()
            fig.savefig(str(out_dir / f"knee_theta_frame_{frame_idx:05d}.png"), dpi=150)
            plt.close(fig)

            mag_in = mag[inside_dir]
            if mag_in.size >= 30:
                dmag_in = np.sort(np.abs(mag_in - mag0))
                fig, ax = plt.subplots(figsize=(6.8, 3.8))
                ax.plot(np.arange(dmag_in.size), dmag_in, linewidth=1.5)
                ax.axhline(r0_mag, linestyle="--", linewidth=1, label=f"r0_mag={r0_mag:.3f}")
                ax.axhline(r_mag, linestyle="--", linewidth=1, label=f"r_mag={r_mag:.3f}")
                ax.set_title(f"Frame {frame_idx} knee on |mag-mag0|")
                ax.set_xlabel("sorted index")
                ax.set_ylabel("|Δmag| (px/frame)")
                ax.grid(True)
                ax.legend(loc="best")
                fig.tight_layout()
                fig.savefig(str(out_dir / f"knee_mag_frame_{frame_idx:05d}.png"), dpi=150)
                plt.close(fig)

            abs_dtheta = dtheta_deg
            fig, ax = plt.subplots(figsize=(6.2, 4.2))
            ax.scatter(abs_dtheta[~target_s], mag[~target_s], s=4, alpha=0.25, label="bg-like")
            ax.scatter(abs_dtheta[target_s], mag[target_s], s=7, alpha=0.75, label="target")
            ax.axvline(r_theta, linestyle="--", linewidth=1, label="theta radius")
            ax.axhline(mag0, linestyle="--", linewidth=1, label="mag0")
            ax.set_xlabel("|Δθ| to dominant dir (deg)")
            ax.set_ylabel("mag (px/frame)")
            ax.set_title(f"Frame {frame_idx} angle vs magnitude")
            ax.grid(True)
            ax.legend(loc="best")
            fig.tight_layout()
            fig.savefig(str(out_dir / f"angle_mag_frame_{frame_idx:05d}.png"), dpi=150)
            plt.close(fig)

            abs_dmag = np.abs(mag - mag0)
            fig, ax = plt.subplots(figsize=(6.2, 4.2))
            ax.scatter(abs_dtheta[~target_s], abs_dmag[~target_s], s=4, alpha=0.25, label="bg-like")
            ax.scatter(abs_dtheta[target_s], abs_dmag[target_s], s=7, alpha=0.75, label="target")
            ax.axvline(r_theta, linestyle="--", linewidth=1, label="theta radius")
            ax.axhline(r_mag, linestyle="--", linewidth=1, label="mag radius")
            ax.set_xlabel("|Δθ| to dominant dir (deg)")
            ax.set_ylabel("|Δmag| = |mag - mag0| (px/frame)")
            ax.set_title(f"Frame {frame_idx} angle deviation vs magnitude deviation")
            ax.grid(True)
            ax.legend(loc="best")
            fig.tight_layout()
            fig.savefig(str(out_dir / f"angle_dmag_frame_{frame_idx:05d}.png"), dpi=150)
            plt.close(fig)

        if vis_cfg["SAVE_DEBUG_TABLE_CSV"]:
            table_path = out_dir / f"table_frame_{frame_idx:05d}.csv"
            with open(table_path, "w", newline="", encoding="utf-8") as f:
                wcsv = csv.writer(f)
                wcsv.writerow([
                    "sx", "sy", "fx", "fy", "mag", "angle_deg",
                    "dom_dir_deg", "abs_dtheta_deg", "r_theta_deg", "inside_dir",
                    "mag0", "abs_dmag", "r_mag", "mag_outlier", "is_target"
                ])
                dom_dir_deg = (np.rad2deg(dom_dir) + 360.0) % 360.0
                angle_deg = (np.rad2deg(theta) + 360.0) % 360.0
                abs_dmag = np.abs(mag - mag0)
                for i in range(pts.shape[0]):
                    wcsv.writerow([
                        int(xs[i]), int(ys[i]),
                        float(pts[i, 0]), float(pts[i, 1]),
                        float(mag[i]), float(angle_deg[i]),
                        float(dom_dir_deg), float(dtheta_deg[i]), float(r_theta), int(inside_dir[i]),
                        float(mag0), float(abs_dmag[i]), float(r_mag), int(mag_outlier[i]), int(target_s[i]),
                    ])
    except Exception as e:
        if logger is not None:
            logger.warning("Debug 保存失败 frame %d: %s", frame_idx, e)
        else:
            print(f"[WARN] Debug 保存失败 frame {frame_idx}: {e}")


def _apply_basic_style(ws, normal_border, center_align):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_align
            cell.border = normal_border


def _set_text_grid_layout(ws):
    ws.column_dimensions["A"].width = 8
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 10
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 22
    ws.freeze_panes = "B2"


def _set_square_like_grid_layout(ws, col_width=4.5, row_height=24):
    ws.column_dimensions["A"].width = 8
    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = col_width
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = row_height
    ws.freeze_panes = "B2"


def save_debug_excel_prepost(debug_pack, post_pack, frame_idx: int, vis_cfg: Dict, paths: Dict, logger=None):
    if (not vis_cfg["SAVE_DEBUG_EXCEL_PRE"]) and (not vis_cfg["SAVE_COMPARE_EXCEL"]):
        return

    try:
        import pandas as pd
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        pts = debug_pack["pts"]
        xs = debug_pack["xs"]
        ys = debug_pack["ys"]
        mag = debug_pack["mag"]
        theta = debug_pack["theta"]
        dom_dir = debug_pack["dom_dir"]
        dtheta_deg = debug_pack["dtheta_deg"]
        r_theta = debug_pack["r_theta"]
        r0_theta = debug_pack["r0_theta"]
        mag0 = debug_pack["mag0"]
        r_mag = debug_pack["r_mag"]
        r0_mag = debug_pack["r0_mag"]
        inside_dir = debug_pack["inside_dir"].astype(bool)
        mag_outlier = debug_pack["mag_outlier"].astype(bool)
        target_s = debug_pack["target_s"].astype(bool)

        angle_deg = (np.rad2deg(theta) + 360.0) % 360.0
        dom_dir_deg = (np.rad2deg(dom_dir) + 360.0) % 360.0
        abs_dmag = np.abs(mag - mag0)

        x_unique = np.sort(np.unique(xs))
        y_unique = np.sort(np.unique(ys))
        out_xlsx = paths["excel_dir"] / f"grid_frame_{frame_idx:05d}.xlsx"

        text_grid = pd.DataFrame("", index=y_unique, columns=x_unique)
        mag_grid = pd.DataFrame(np.nan, index=y_unique, columns=x_unique)
        angle_grid = pd.DataFrame(np.nan, index=y_unique, columns=x_unique)
        dtheta_grid = pd.DataFrame(np.nan, index=y_unique, columns=x_unique)
        dmag_grid = pd.DataFrame(np.nan, index=y_unique, columns=x_unique)
        target_grid_pre = pd.DataFrame(0, index=y_unique, columns=x_unique)
        inside_dir_grid = pd.DataFrame(0, index=y_unique, columns=x_unique)
        target_grid_post = pd.DataFrame(0, index=y_unique, columns=x_unique)
        target_grid_morph = pd.DataFrame(0, index=y_unique, columns=x_unique)
        target_diff_grid = pd.DataFrame(0, index=y_unique, columns=x_unique)
        region_id_grid_post = pd.DataFrame(0, index=y_unique, columns=x_unique)

        post_mask_bool = post_pack["post_mask_bool"]
        morph_mask_bool = post_pack["morph_mask_bool"]
        region_id_map = post_pack["region_id_map"]

        for i in range(len(xs)):
            x = int(xs[i])
            y = int(ys[i])
            if vis_cfg["EXCEL_TEXT_SHOW_ANGLE_MAG"]:
                text_grid.loc[y, x] = f"{angle_deg[i]:.0f}|{mag[i]:.2f}"
            mag_grid.loc[y, x] = float(mag[i])
            angle_grid.loc[y, x] = float(angle_deg[i])
            dtheta_grid.loc[y, x] = float(dtheta_deg[i])
            dmag_grid.loc[y, x] = float(abs_dmag[i])
            target_grid_pre.loc[y, x] = 1 if target_s[i] else 0
            inside_dir_grid.loc[y, x] = 1 if inside_dir[i] else 0

            post_v = 1 if post_mask_bool[y, x] else 0
            morph_v = 1 if morph_mask_bool[y, x] else 0
            pre_v = 1 if target_s[i] else 0
            region_id_v = int(region_id_map[y, x]) if region_id_map is not None else 0

            target_grid_post.loc[y, x] = post_v
            target_grid_morph.loc[y, x] = morph_v
            region_id_grid_post.loc[y, x] = region_id_v

            if pre_v == 0 and post_v == 0:
                diff_v = 0
            elif pre_v == 1 and post_v == 0:
                diff_v = 1
            elif pre_v == 1 and post_v == 1:
                diff_v = 2
            else:
                diff_v = 3
            target_diff_grid.loc[y, x] = diff_v

        summary_pre_df = pd.DataFrame([
            ["frame_idx", frame_idx],
            ["dom_dir_deg", dom_dir_deg],
            ["r_theta_deg", r_theta],
            ["r0_theta_deg", r0_theta],
            ["mag0", mag0],
            ["r_mag", r_mag],
            ["r0_mag", r0_mag],
            ["num_points", len(xs)],
            ["num_target_points_pre_sample", int(target_s.sum())],
            ["num_inside_dir_points", int(inside_dir.sum())],
            ["num_mag_outlier_points", int(mag_outlier.sum())],
        ], columns=["item", "value"])

        summary_post_df = pd.DataFrame([
            ["frame_idx", frame_idx],
            ["pre_target_pixels_full", int(post_pack["pre_target_pixels"])],
            ["morph_target_pixels_full", int(post_pack["morph_target_pixels"])],
            ["post_target_pixels_full", int(post_pack["post_target_pixels"])],
            ["num_raw_contours", int(post_pack["num_raw_contours"])],
            ["num_kept_contours", int(post_pack["num_kept_contours"])],
            ["removed_sample_points", int((target_diff_grid.values == 1).sum())],
            ["kept_sample_points", int((target_diff_grid.values == 2).sum())],
            ["added_sample_points", int((target_diff_grid.values == 3).sum())],
        ], columns=["item", "value"])

        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            # ===== 1. 概要信息 =====
            summary_pre_df.to_excel(writer, sheet_name="summary_pre", index=False)
            summary_post_df.to_excel(writer, sheet_name="summary_post", index=False)

            # ===== 2. 最直观的总览表 =====
            if vis_cfg["SAVE_DEBUG_EXCEL_PRE"]:
                if vis_cfg["EXCEL_SAVE_TEXT_GRID"]:
                    text_grid.to_excel(writer, sheet_name="text_grid_pre")

            # ===== 3. 检测结果流程：pre → morph → post → diff → region =====
            if vis_cfg["SAVE_DEBUG_EXCEL_PRE"]:
                if vis_cfg["EXCEL_SAVE_TARGET_GRID"]:
                    target_grid_pre.to_excel(writer, sheet_name="target_grid_pre")

            if vis_cfg["SAVE_COMPARE_EXCEL"]:
                target_grid_morph.to_excel(writer, sheet_name="target_grid_morph")
                target_grid_post.to_excel(writer, sheet_name="target_grid_post")
                target_diff_grid.to_excel(writer, sheet_name="target_diff_grid")
                region_id_grid_post.to_excel(writer, sheet_name="region_id_grid_post")

            # ===== 4. 判断依据：angle → dtheta → inside_dir → mag → dmag =====
            if vis_cfg["SAVE_DEBUG_EXCEL_PRE"]:
                if vis_cfg["EXCEL_SAVE_ANGLE_GRID"]:
                    angle_grid.to_excel(writer, sheet_name="angle_grid_pre")
                if vis_cfg["EXCEL_SAVE_DTHETA_GRID"]:
                    dtheta_grid.to_excel(writer, sheet_name="dtheta_grid_pre")
                if vis_cfg["EXCEL_SAVE_INSIDE_DIR_GRID"]:
                    inside_dir_grid.to_excel(writer, sheet_name="inside_dir_grid_pre")
                if vis_cfg["EXCEL_SAVE_MAG_GRID"]:
                    mag_grid.to_excel(writer, sheet_name="mag_grid_pre")
                if vis_cfg["EXCEL_SAVE_DMAG_GRID"]:
                    dmag_grid.to_excel(writer, sheet_name="dmag_grid_pre")
        #
        # with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        #     summary_pre_df.to_excel(writer, sheet_name="summary_pre", index=False)
        #     summary_post_df.to_excel(writer, sheet_name="summary_post", index=False)
        #
        #     if vis_cfg["SAVE_DEBUG_EXCEL_PRE"]:
        #         if vis_cfg["EXCEL_SAVE_TEXT_GRID"]:
        #             text_grid.to_excel(writer, sheet_name="text_grid_pre")
        #         if vis_cfg["EXCEL_SAVE_MAG_GRID"]:
        #             mag_grid.to_excel(writer, sheet_name="mag_grid_pre")
        #         if vis_cfg["EXCEL_SAVE_ANGLE_GRID"]:
        #             angle_grid.to_excel(writer, sheet_name="angle_grid_pre")
        #         if vis_cfg["EXCEL_SAVE_DTHETA_GRID"]:
        #             dtheta_grid.to_excel(writer, sheet_name="dtheta_grid_pre")
        #         if vis_cfg["EXCEL_SAVE_DMAG_GRID"]:
        #             dmag_grid.to_excel(writer, sheet_name="dmag_grid_pre")
        #         if vis_cfg["EXCEL_SAVE_TARGET_GRID"]:
        #             target_grid_pre.to_excel(writer, sheet_name="target_grid_pre")
        #         if vis_cfg["EXCEL_SAVE_INSIDE_DIR_GRID"]:
        #             inside_dir_grid.to_excel(writer, sheet_name="inside_dir_grid_pre")
        #
        #     if vis_cfg["SAVE_COMPARE_EXCEL"]:
        #         target_grid_morph.to_excel(writer, sheet_name="target_grid_morph")
        #         target_grid_post.to_excel(writer, sheet_name="target_grid_post")
        #         target_diff_grid.to_excel(writer, sheet_name="target_diff_grid")
        #         region_id_grid_post.to_excel(writer, sheet_name="region_id_grid_post")

            wb = writer.book
            thin_gray = Side(style="thin", color="CCCCCC")
            red_side = Side(style="medium", color="FF0000")
            blue_side = Side(style="medium", color="0070C0")
            green_side = Side(style="medium", color="00B050")
            red_border = Border(left=red_side, right=red_side, top=red_side, bottom=red_side)
            blue_border = Border(left=blue_side, right=blue_side, top=blue_side, bottom=blue_side)
            green_border = Border(left=green_side, right=green_side, top=green_side, bottom=green_side)
            normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
            red_fill = PatternFill(fill_type="solid", fgColor="FFD9D9")
            yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
            gray_fill = PatternFill(fill_type="solid", fgColor="FFF2F2F2")
            green_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
            blue_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
            bold_font = Font(bold=True)
            red_font = Font(color="C00000", bold=True)
            center_align = Alignment(horizontal="center", vertical="center")

            for sname in ["summary_pre", "summary_post"]:
                ws = wb[sname]
                _apply_basic_style(ws, normal_border, center_align)
                ws.column_dimensions["A"].width = 24
                ws.column_dimensions["B"].width = 18
                ws.freeze_panes = "A2"

            grid_sheets = [
                "text_grid_pre", "mag_grid_pre", "angle_grid_pre", "dtheta_grid_pre", "dmag_grid_pre",
                "target_grid_pre", "inside_dir_grid_pre", "target_grid_morph", "target_grid_post",
                "target_diff_grid", "region_id_grid_post",
            ]
            for sname in grid_sheets:
                if sname in wb.sheetnames:
                    _apply_basic_style(wb[sname], normal_border, center_align)

            if "text_grid_pre" in wb.sheetnames:
                ws_txt = wb["text_grid_pre"]
                _set_text_grid_layout(ws_txt)
                x_to_col = {int(x): i + 2 for i, x in enumerate(x_unique)}
                y_to_row = {int(y): i + 2 for i, y in enumerate(y_unique)}
                for i in range(len(xs)):
                    r = y_to_row[int(ys[i])]
                    c = x_to_col[int(xs[i])]
                    cell = ws_txt.cell(r, c)
                    if vis_cfg["EXCEL_ENABLE_COLOR"]:
                        val = float(dtheta_deg[i])
                        if val <= r_theta * 0.5:
                            cell.fill = gray_fill
                        elif val <= r_theta:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill
                    if vis_cfg["EXCEL_HIGHLIGHT_TARGET"] and target_s[i]:
                        cell.border = red_border
                        cell.font = red_font
                    elif vis_cfg["EXCEL_HIGHLIGHT_INSIDE_DIR"] and inside_dir[i]:
                        cell.border = blue_border

            for sname in [
                "mag_grid_pre", "angle_grid_pre", "dtheta_grid_pre", "dmag_grid_pre", "target_grid_pre",
                "inside_dir_grid_pre", "target_grid_morph", "target_grid_post", "target_diff_grid", "region_id_grid_post",
            ]:
                if sname in wb.sheetnames:
                    _set_square_like_grid_layout(wb[sname], col_width=4.5, row_height=24)

            if vis_cfg["EXCEL_ENABLE_COLOR"]:
                for sname in ["mag_grid_pre", "dtheta_grid_pre", "dmag_grid_pre", "angle_grid_pre", "region_id_grid_post"]:
                    if sname not in wb.sheetnames:
                        continue
                    ws = wb[sname]
                    max_row = ws.max_row
                    max_col = ws.max_column
                    if max_row >= 2 and max_col >= 2:
                        rng = f"B2:{ws.cell(max_row, max_col).coordinate}"
                        ws.conditional_formatting.add(
                            rng,
                            ColorScaleRule(
                                start_type="min", start_color="FFF2F2F2",
                                mid_type="percentile", mid_value=50, mid_color="FF9DC3E6",
                                end_type="max", end_color="FF2F75B5",
                            ),
                        )

            if "target_grid_pre" in wb.sheetnames and vis_cfg["EXCEL_HIGHLIGHT_TARGET"]:
                ws = wb["target_grid_pre"]
                for row in ws.iter_rows(min_row=2, min_col=2):
                    for cell in row:
                        if cell.value == 1:
                            if vis_cfg["EXCEL_ENABLE_COLOR"]:
                                cell.fill = red_fill
                            cell.font = bold_font
                            cell.border = red_border

            if "inside_dir_grid_pre" in wb.sheetnames and vis_cfg["EXCEL_HIGHLIGHT_INSIDE_DIR"]:
                ws = wb["inside_dir_grid_pre"]
                for row in ws.iter_rows(min_row=2, min_col=2):
                    for cell in row:
                        if cell.value == 1:
                            cell.border = blue_border

            if "target_grid_morph" in wb.sheetnames:
                ws = wb["target_grid_morph"]
                for row in ws.iter_rows(min_row=2, min_col=2):
                    for cell in row:
                        if cell.value == 1:
                            if vis_cfg["EXCEL_ENABLE_COLOR"]:
                                cell.fill = yellow_fill
                            cell.border = green_border

            if "target_grid_post" in wb.sheetnames:
                ws = wb["target_grid_post"]
                for row in ws.iter_rows(min_row=2, min_col=2):
                    for cell in row:
                        if cell.value == 1:
                            if vis_cfg["EXCEL_ENABLE_COLOR"]:
                                cell.fill = green_fill
                            cell.font = bold_font
                            cell.border = green_border

            if "target_diff_grid" in wb.sheetnames:
                ws = wb["target_diff_grid"]
                for row in ws.iter_rows(min_row=2, min_col=2):
                    for cell in row:
                        v = int(cell.value) if cell.value is not None else 0
                        if v == 1:
                            cell.fill = red_fill
                            cell.border = red_border
                        elif v == 2:
                            cell.fill = green_fill
                            cell.border = green_border
                        elif v == 3:
                            cell.fill = blue_fill
                            cell.border = blue_border
        if logger is not None:
            logger.info("Excel 网格表已保存: %s", out_xlsx)
        else:
            print("[OK] Excel 网格表已保存：", out_xlsx)
    except Exception as e:
        if logger is not None:
            logger.warning("Excel 保存失败 frame %d: %s", frame_idx, e)
        else:
            print(f"[WARN] Excel 保存失败 frame {frame_idx}: {e}")


def save_compare_outputs(frame, frame_idx: int, post_pack: Dict, region_frame, box_frame, combined_frame, vis_cfg: Dict, paths: Dict, logger=None):
    pre_mask_bool = post_pack["pre_mask_bool"]
    post_mask_bool = post_pack["post_mask_bool"]

    if vis_cfg["SAVE_COMPARE_PRE_MASK_IMAGE"]:
        safe_imwrite(paths["compare_pre_dir"] / f"mask_pre_frame_{frame_idx:05d}.png", post_pack["pre_mask_u8"], logger=logger)

    if vis_cfg["SAVE_COMPARE_PRE_OVERLAY_IMAGE"]:
        overlay_pre = overlay_mask(frame, pre_mask_bool, color=(0, 0, 255), alpha=0.35)
        safe_imwrite(paths["compare_pre_dir"] / f"overlay_pre_frame_{frame_idx:05d}.png", overlay_pre, logger=logger)

    if vis_cfg["SAVE_COMPARE_POST_MORPH_MASK_IMAGE"]:
        safe_imwrite(paths["compare_post_dir"] / f"mask_post_morph_frame_{frame_idx:05d}.png", post_pack["morph_mask_u8"], logger=logger)

    if vis_cfg["SAVE_COMPARE_POST_FINAL_MASK_IMAGE"]:
        safe_imwrite(paths["compare_post_dir"] / f"mask_post_final_frame_{frame_idx:05d}.png", post_pack["post_mask_u8"], logger=logger)

    if vis_cfg["SAVE_COMPARE_POST_OVERLAY_IMAGE"]:
        overlay_post = overlay_mask(frame, post_mask_bool, color=(0, 255, 0), alpha=0.35)
        safe_imwrite(paths["compare_post_dir"] / f"overlay_post_frame_{frame_idx:05d}.png", overlay_post, logger=logger)

    if vis_cfg["SAVE_COMPARE_REGION_IMAGE"]:
        safe_imwrite(paths["compare_post_dir"] / f"region_post_frame_{frame_idx:05d}.png", region_frame, logger=logger)

    if vis_cfg["SAVE_COMPARE_BOX_IMAGE"]:
        safe_imwrite(paths["compare_post_dir"] / f"box_post_frame_{frame_idx:05d}.png", box_frame, logger=logger)

    if vis_cfg["SAVE_COMPARE_COMBINED_IMAGE"]:
        safe_imwrite(paths["compare_post_dir"] / f"combined_post_frame_{frame_idx:05d}.png", combined_frame, logger=logger)


def save_raw_speed_hist(all_mags, paths: Dict, vis_cfg: Dict, logger=None):
    if len(all_mags) == 0:
        return
    all_mags = np.concatenate(all_mags, axis=0)
    all_mags = all_mags[np.isfinite(all_mags)]
    if all_mags.size == 0:
        return

    max_mag = float(all_mags.max()) if all_mags.max() > 0 else 1.0
    bins = np.linspace(0, max_mag, 201)
    hist, edges = np.histogram(all_mags, bins=bins)
    centers_mag = 0.5 * (edges[:-1] + edges[1:])

    if vis_cfg["SAVE_RAW_SPEED_HIST_CSV"]:
        with open(paths["raw_speed_hist_csv"], "w", newline="", encoding="utf-8") as fh:
            hw = csv.writer(fh)
            hw.writerow(["speed_bin_center_px_per_frame", "pixel_count"])
            for s, c in zip(centers_mag, hist):
                hw.writerow([f"{s:.6f}", int(c)])

    if vis_cfg["SAVE_RAW_SPEED_HIST_PNG"]:
        try:
            import matplotlib.pyplot as plt
            plt.figure(figsize=(8, 4))
            plt.plot(centers_mag, hist)
            plt.xlabel("Speed (pixels/frame)")
            plt.ylabel("Number of pixels")
            plt.title("Distribution of optical flow magnitude (RAW)")
            plt.tight_layout()
            plt.savefig(paths["raw_speed_hist_png"], dpi=150)
            plt.close()
            print("[OK] 原始速度分布图保存：", paths["raw_speed_hist_png"])
        except Exception as e:
            print("[WARN] 无法绘制原始速度分布 PNG，仅保存 CSV。错误信息：", e)
